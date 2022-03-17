import openpyxl

# load excel with its path
wrkbk = openpyxl.load_workbook("Book1.xlsx")

sh = wrkbk.active

# iterate through excel and display data
for i in range(2, sh.max_row + 1):
    print("\n")
    print("Row ", i, " data :")

    for j in range(1, sh.max_column + 1):
        print(sh.cell(row=i, column=j).value,end=" ")
        #print(cell_obj.value, end=" ")