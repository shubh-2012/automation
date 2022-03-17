

import tkinter as tk
import win32com.client
import sys
import subprocess
import time
import openpyxl
from pathlib import Path

xlsx_file = Path('zxc.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active
username = sheet["A1"].value
password = sheet["B1"].value
package  = sheet["C1"].value
wrkbk = openpyxl.load_workbook("Book1.xlsx")

sheet2 = wrkbk.active

def sap():
    try:

        path = r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplogon.exe"
        subprocess.Popen(path)
        time.sleep(10)

        SapGuiAuto = win32com.client.GetObject('SAPGUI')
        if not type(SapGuiAuto) == win32com.client.CDispatch:
            return
        application = SapGuiAuto.GetScriptingEngine
        if not type(application) == win32com.client.CDispatch:
            SapGuiAuto = None
            return
        connection = application.OpenConnection('S48 [IN-BLR-1709]', True)

        if not type(connection) == win32com.client.CDispatch:
            application = None
            SapGuiAuto = None
            return
        session = connection.Children(0)
        if not type(session) == win32com.client.CDispatch:
            connection = None
            application = None
            SapGuiAuto = None
            return
        session.findById("wnd[0]").maximize()

        session.findById("wnd[0]/usr/txtRSYST-MANDT").text = 600
        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = username
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = password
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").setFocus()
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").caretPosition = 9
        session.findById("wnd[0]").sendVKey(0)

        session.findById("wnd[1]/usr/radMULTI_LOGON_OPT2").select()
        session.findById("wnd[1]/usr/radMULTI_LOGON_OPT2").setFocus()
        session.findById("wnd[1]/tbar[0]/btn[0]").press()

        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").text = "se11"
        session.findById("wnd[0]").sendVKey(0)

        for i in range(2, sheet2.max_row + 1):

            j = 1

            session.findById("wnd[0]/usr/radRSRD1-DDTYPE").setFocus()
            session.findById("wnd[0]/usr/radRSRD1-DDTYPE").select()
            session.findById("wnd[0]/usr/ctxtRSRD1-DDTYPE_VAL").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/ctxtRSRD1-DDTYPE_VAL").caretPosition = 13
            session.findById("wnd[0]/usr/btnPUSHADD").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/usr/txtDD04D-DDTEXT").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTS/tabpTYPE/ssubSUB_DATA:SAPLSD51:1002/ctxtDD04D-DOMNAME").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTS/tabpTYPE/ssubSUB_DATA:SAPLSD51:1002/ctxtDD04D-DOMNAME").setFocus()
            session.findById("wnd[0]/usr/tabsTS/tabpTYPE/ssubSUB_DATA:SAPLSD51:1002/ctxtDD04D-DOMNAME").caretPosition = 9
            session.findById("wnd[0]").sendVKey(2)
            session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
            session.findById("wnd[1]/usr/ctxtKO007-L_DEVCLASS").text = package
            session.findById("wnd[1]/usr/ctxtKO007-L_DEVCLASS").caretPosition = 12
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[1]/usr/btnSPOP-OPTION1").press()
            session.findById("wnd[0]/usr/txtDD01D-DDTEXT").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/ctxtDD01D-DATATYPE").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/txtDD01D-LENG").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/txtDD01D-DECIMALS").text = sheet2.cell(row=i, column=j).value
            j += 1
            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/txtDD01D-OUTPUTLEN").text = sheet2.cell(row=i, column=j).value

            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/txtDD01D-OUTPUTLEN").setFocus()
            session.findById("wnd[0]/usr/tabsTAB_STRIP/tabpTAB1/ssubTS_SCREEN:SAPLSD11:1201/txtDD01D-OUTPUTLEN").caretPosition = 6
            session.findById("wnd[0]").sendVKey(11)
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/tbar[1]/btn[27]").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/tbar[1]/btn[18]").press()
            session.findById("wnd[0]/tbar[1]/btn[27]").press()
            session.findById("wnd[1]/tbar[0]/btn[0]").press()
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nse11"
            session.findById("wnd[0]").sendVKey(0)



    except Exception as e:
        print("Error in Execution",e)



    finally:
        session = None
        connection = None
        application = None
        SapGuiAuto = None

sap()
